from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Palette ──────────────────────────────────────────────────────────────────
HEADER_BG   = "1F3864"   # bleu foncé
HEADER_FG   = "FFFFFF"
SECTION_BG  = "2E75B6"   # bleu moyen
SECTION_FG  = "FFFFFF"
ROW_ODD     = "EBF3FB"
ROW_EVEN    = "FFFFFF"
FONT_NAME   = "Arial"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_style(cell, text):
    cell.value = text
    cell.font = Font(name=FONT_NAME, bold=True, color=HEADER_FG, size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

def section_style(cell, text):
    cell.value = text
    cell.font = Font(name=FONT_NAME, bold=True, color=SECTION_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=SECTION_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border

def data_cell(cell, text, row_idx):
    cell.value = text
    cell.font = Font(name=FONT_NAME, size=10)
    cell.fill = PatternFill("solid", fgColor=ROW_ODD if row_idx % 2 == 0 else ROW_EVEN)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Base de données
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Base de données"
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 22

cols = ["Projet", "Host", "Port", "Base de données", "Utilisateur", "Mot de passe", "DATABASE_URL"]
for c, label in enumerate(cols, 1):
    header_style(ws1.cell(1, c), label)

rows_db = [
    ("LeComptoirImmo", "localhost", "5432", "lecomptoirimmo", "lecomptoirimmo_user",
     "devpassword123",
     "postgresql+asyncpg://lecomptoirimmo_user:devpassword123@localhost:5432/lecomptoirimmo"),
    ("Alice", "localhost", "5432", "lecomptoirimmo", "lecomptoirimmo_user",
     "devpassword123",
     "postgresql+asyncpg://lecomptoirimmo_user:devpassword123@localhost:5432/lecomptoirimmo"),
]

for r, row in enumerate(rows_db, 2):
    ws1.row_dimensions[r].height = 20
    for c, val in enumerate(row, 1):
        data_cell(ws1.cell(r, c), val, r)

set_col_widths(ws1, [20, 14, 8, 20, 24, 18, 72])

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Comptes applicatifs
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Comptes applicatifs")
ws2.row_dimensions[1].height = 30

cols2 = ["Projet", "Rôle", "Email", "Mot de passe", "Notes"]
for c, label in enumerate(cols2, 1):
    header_style(ws2.cell(1, c), label)

accounts = [
    # LeComptoirImmo
    ("LeComptoirImmo", "Admin",                   "admin@locataire-cloud.fr",          "Admin1234!",            "Compte administrateur principal"),
    ("LeComptoirImmo", "Gestionnaire mandataire",  "gestionnaire@cabinet.fr",           "Gestionnaire1!",        "Compte test gestionnaire"),
    ("LeComptoirImmo", "Gestionnaire-propriétaire","gestionnaire-proprio@cabinet.fr",   "GestionnaireProprio1!", "Compte test GP"),
    ("LeComptoirImmo", "Gestionnaire-propriétaire","residence.tatie@outlook.com",       "—",                     "Compte GP secondaire"),
    ("LeComptoirImmo", "Propriétaire",             "proprietaire@email.fr",             "Proprietaire1!",        "Compte test propriétaire"),
    ("LeComptoirImmo", "Locataire",                "locataire@email.fr",                "Locataire1!",           "Compte test locataire"),
    ("LeComptoirImmo", "Locataire (GP)",           "t@gmail.com",                       "—",                     "Créé par gestionnaire-proprio"),
    ("LeComptoirImmo", "Locataire (GP)",           "hdhd@gmail.com",                    "—",                     "Créé par gestionnaire-proprio"),
    ("LeComptoirImmo", "Locataire (GP)",           "test.isolation@test.fr",            "—",                     "Créé par gestionnaire-proprio"),
    # Alice
    ("Alice",       "Admin",                   "admin@alice.fr",                 "Alice1!",            "Compte administrateur Alice"),
]

for r, row in enumerate(accounts, 2):
    ws2.row_dimensions[r].height = 20
    for c, val in enumerate(row, 1):
        data_cell(ws2.cell(r, c), val, r)

set_col_widths(ws2, [20, 26, 38, 22, 34])

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Configuration applicative
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Configuration")
ws3.row_dimensions[1].height = 30

cols3 = ["Projet", "Clé", "Valeur", "Description"]
for c, label in enumerate(cols3, 1):
    header_style(ws3.cell(1, c), label)

configs = [
    ("LeComptoirImmo", "SECRET_KEY",                   "dev-secret-key-change-in-production-openssl-rand-hex-32", "JWT — À changer en prod"),
    ("LeComptoirImmo", "ALGORITHM",                    "HS256",                                                    "Algorithme JWT"),
    ("LeComptoirImmo", "ACCESS_TOKEN_EXPIRE_MINUTES",  "30",                                                       "Durée token accès (minutes)"),
    ("LeComptoirImmo", "REFRESH_TOKEN_EXPIRE_DAYS",    "7",                                                        "Durée token refresh (jours)"),
    ("LeComptoirImmo", "BACKEND_URL",                  "http://localhost:8000",                                     "URL backend FastAPI"),
    ("LeComptoirImmo", "FRONTEND_URL",                 "http://localhost:5175",                                     "URL frontend Vite (port variable)"),
    ("LeComptoirImmo", "CORS_ORIGINS",                 "http://localhost:5173 → 5177, :3000",                      "Ports Vite autorisés"),
    ("Alice",       "SECRET_KEY",                   "alice-dev-secret-key-change-in-production",             "JWT Alice — À changer en prod"),
    ("Alice",       "ACCESS_TOKEN_EXPIRE_MINUTES",  "60",                                                       "Durée token accès (minutes)"),
    ("Alice",       "CORS_ORIGINS",                 "http://localhost:5174",                                     "Port frontend Alice"),
]

for r, row in enumerate(configs, 2):
    ws3.row_dimensions[r].height = 20
    for c, val in enumerate(row, 1):
        data_cell(ws3.cell(r, c), val, r)

set_col_widths(ws3, [20, 34, 54, 36])

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — GitHub & SSH
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("GitHub & SSH")
ws4.row_dimensions[1].height = 30

cols4 = ["Élément", "Valeur", "Notes"]
for c, label in enumerate(cols4, 1):
    header_style(ws4.cell(1, c), label)

github_rows = [
    ("Compte GitHub",            "service-lecomptoir",                                       "Organisation GitHub"),
    ("Repo LeComptoirImmo",      "git@github.com:service-lecomptoir/LeComptoirImmo.git",     "Remote SSH"),
    ("Repo LeComptoirMarket",    "git@github.com:service-lecomptoir/LeComptoirMarket.git",   "Remote SSH"),
    ("Clé SSH — nom",            "lecomptoir",                                               "Partagée entre tous les repos"),
    ("Clé SSH — privée",         "C:\\Users\\Arons\\.ssh\\lecomptoir",                       "Ne jamais partager"),
    ("Clé SSH — publique",       "C:\\Users\\Arons\\.ssh\\lecomptoir.pub",                   "Copiée dans GitHub Settings > SSH keys"),
    ("SSH config",               "C:\\Users\\Arons\\.ssh\\config",                           "IdentityFile → ~/.ssh/lecomptoir"),
    ("Empreinte clé (SHA256)",   "SHA256:J0HkwMjEX0ysynAX1vjBfu9DtC8rviKkvSgh+DZD8qk",     "Ed25519"),
]

for r, row in enumerate(github_rows, 2):
    ws4.row_dimensions[r].height = 20
    for c, val in enumerate(row, 1):
        data_cell(ws4.cell(r, c), val, r)

set_col_widths(ws4, [28, 62, 36])

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Ports & services
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Ports & Services")
ws5.row_dimensions[1].height = 30

cols5 = ["Service", "Port", "URL", "Notes"]
for c, label in enumerate(cols5, 1):
    header_style(ws5.cell(1, c), label)

ports = [
    ("Backend FastAPI (LeComptoirImmo)", "8000", "http://localhost:8000",      "uvicorn app.main:app"),
    ("Frontend Vite (LeComptoirImmo)",   "5173–5177", "http://localhost:517x", "Port auto-assigné par Vite"),
    ("PostgreSQL",                        "5432", "localhost:5432",             "Base lecomptoirimmo"),
    ("API Docs (Swagger)",               "8000", "http://localhost:8000/docs",  "Documentation interactive"),
    ("Health check",                     "8000", "http://localhost:8000/health","Vérification backend"),
]

for r, row in enumerate(ports, 2):
    ws5.row_dimensions[r].height = 20
    for c, val in enumerate(row, 1):
        data_cell(ws5.cell(r, c), val, r)

set_col_widths(ws5, [36, 14, 36, 30])

# ── Sauvegarde ───────────────────────────────────────────────────────────────
out = r"C:\Users\Arons\OneDrive\Documents\Work Place\Claude IA\LeComptoirImmo\credentials_lecomptoir.xlsx"
wb.save(out)
print(f"Fichier créé : {out}")
